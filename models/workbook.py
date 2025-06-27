from models.projeto import Dia, Semana, Projeto

from openpyxl import load_workbook
from datetime import timezone, date, timedelta
import shutil
import os
import streamlit as st

class WorkbookManager:
    def __init__(self, filename):
        self.original_filename = filename
        self.copied_filename = self.create_copy(filename)
        self.workbook = load_workbook(self.copied_filename)
        self.cronograma_geral = CronogramaGeral(self.workbook)
        self.proximas_semanas = ProximasSemanas(self.workbook)
        self.planilhas_semanais = PlanilhasSemanais(self.workbook)

    def create_copy(self, filename):
        new_filename = os.path.join(os.path.dirname(filename), "Master Planejamento Atualizado.xlsx")

        # Só copia se o nome de destino for diferente do original
        if os.path.abspath(filename) != os.path.abspath(new_filename):
            shutil.copy(filename, new_filename)

        return os.path.abspath(new_filename)

    def save(self):
        self.workbook.save(self.copied_filename)

    def close(self):
        pass


class CronogramaGeral:

    def __init__(self, workbook):
        self.sheet = workbook['Crono. Geral']
        self.lin_datas = 6
        self.lin_inicio_lista_projetos = 7
        self.col_inicio_datas = 10
        self.col_nomes_projetos = 2
        self.lin_feriado = 5

    def checa_feriado(self, col):
        valor = self.sheet.cell(row=self.lin_feriado, column=col).value
        return valor == "FERIADO"

    def get_data_inicio(self, data_inicio_coleta, dias_pre):
        col = self.get_coluna_inicio(data_inicio_coleta)
        dias = dias_pre
        while dias > 0 and col > 1:
            if not self.checa_feriado(col):
                dias -= 1
            col -= 1
        data = self.sheet.cell(row=self.lin_datas, column=col).value
        return data.date() if hasattr(data, 'date') else data

    def get_coluna_inicio(self, data_inicio):
        col = self.col_inicio_datas
        max_col = self.sheet.max_column
        while col <= max_col:
            val = self.sheet.cell(row=self.lin_datas, column=col).value
            if val is None:
                col += 1
                continue
            val_date = val.date() if hasattr(val, 'date') else val
            if val_date == data_inicio:
                return col
            col += 1
        return None

    def get_data_fim(self, data_inicio, dias_uteis):
        col = self.get_coluna_inicio(data_inicio)
        if col is None:
            print("Data de início não encontrada.")
            return None
        i = 0
        while i < dias_uteis and col <= self.sheet.max_column:
            if not self.checa_feriado(col):
                i += 1
            col += 1
        val = self.sheet.cell(row=self.lin_datas, column=col - 1).value
        return val.date() if hasattr(val, 'date') else val

    def inserir_projeto(self, specifications, projeto):
        data_inicio = self.get_data_inicio(projeto.data_inicio_coleta, 5)
        col_inicio_proj = self.get_coluna_inicio(data_inicio)
        if col_inicio_proj is None:
            print("Coluna de início não encontrada.")
            return
        row = self.lin_inicio_lista_projetos
        while self.sheet.cell(row=row, column=col_inicio_proj).value:
            row += 1
        self.sheet.insert_rows(row)
        self.sheet.cell(row=row, column=self.col_nomes_projetos, value=projeto.nome)
        col = col_inicio_proj
        for text, count in specifications:
            while count > 0 and col <= self.sheet.max_column:
                if not self.checa_feriado(col):
                    self.sheet.cell(row=row, column=col, value=text)
                    count -= 1
                col += 1
    

class ProximasSemanas:
    
    def __init__(self, workbook):
        self.sheet = workbook['Próximas Semanas']
        self.lin_inicio_lista_projetos = 6
        self.col_PB = 3
        self.col_data_inicio = 6
        self.col_data_fim = 7
        self.col_amostra = 8
        self.col_coleta = 9
        self.col_dur = 11
        self.col_media = 12
        self.col_tma = 13
        self.col_nomes = 16

    def get_linha_projeto(self, data_inicio_coleta):
        for row in range(self.lin_inicio_lista_projetos, self.sheet.max_row + 1):
            val = self.sheet.cell(row=row, column=self.col_data_inicio).value
            if hasattr(val, 'date'):
                val = val.date()
            if val and val >= data_inicio_coleta:
                return row
        return self.sheet.max_row + 1

    def get_diff_producao_pessoas(self):
        linha_producao = None
        linha_pessoas = None

        for row in range(1, self.sheet.max_row + 1):
            val = self.sheet.cell(row=row, column=self.col_nomes).value
            if val == "Produção":
                linha_producao = row
            elif val == "Pessoas":
                linha_pessoas = row
            if linha_producao and linha_pessoas:
                break

        return abs(linha_pessoas - linha_producao) if linha_producao and linha_pessoas else 5

    def get_linha_pessoas(self, data_inicio_coleta):
        return self.get_linha_projeto(data_inicio_coleta) + self.get_diff_producao_pessoas()

    def inserir_projeto_prod(self, projeto):
        media = 1
        tma = 1
        linha = self.get_linha_projeto(projeto.data_inicio_coleta)

        self.sheet.insert_rows(linha)
        self.sheet.cell(row=linha, column=self.col_nomes, value=projeto.nome)
        self.sheet.cell(row=linha, column=self.col_data_inicio, value=projeto.data_inicio_coleta)
        self.sheet.cell(row=linha, column=self.col_data_fim, value=projeto.data_fim)
        self.sheet.cell(row=linha, column=self.col_amostra, value=projeto.amostra)
        self.sheet.cell(row=linha, column=self.col_dur, value=projeto.duracao_coleta)
        self.sheet.cell(row=linha, column=self.col_media, value=media)
        self.sheet.cell(row=linha, column=self.col_tma, value=tma)

    def inserir_projeto_pessoas(self, projeto):
        linha = self.get_linha_pessoas(projeto.data_inicio_coleta)
        self.sheet.insert_rows(linha)
        self.sheet.cell(row=linha, column=self.col_nomes, value=projeto.nome)

    def inserir_projeto(self, projeto):
        self.inserir_projeto_prod(projeto)
        self.inserir_projeto_pessoas(projeto)
        

class PlanilhasSemanais:
    
    def __init__(self, workbook):
        self.workbook = workbook
        self.col_nomes = 5
        self.col_dias_inicio = 6
        self.col_coleta_real = 8
        self.col_hc_real = 11
        self.col_coleta_total = 15
        self.col_prod = 16
        self.col_hc_total = 17
        self.col_hc_parcial = 20
        self.col_dia = self.get_coluna_diaria()
        self.lin_nomes_inicio = 8

    def get_sheets(self, projeto):
        sheets = []
        for semana in projeto.semanas:
            data_str = semana.data_inicio.strftime('%d.%m')
            sheetname = "CATI_Semana_" + data_str
            sheets.append(self.workbook[sheetname])
        return sheets

    def get_lin_monitoramento(self, sheet):
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=self.col_nomes).value == "Monitoramento":
                return row + 3
        return self.lin_nomes_inicio

    def inserir_projeto(self, projeto):
        sheets = self.get_sheets(projeto)
        for sheet, semana in zip(sheets, projeto.semanas):
            linha = self.lin_nomes_inicio
            while sheet.cell(row=linha, column=self.col_nomes).value:
                linha += 1

            sheet.insert_rows(linha)
            sheet.cell(row=linha, column=self.col_nomes, value=projeto.nome)
            for i, dia in enumerate(semana.dias_uteis):
                if not dia.feriado:
                    sheet.cell(row=linha, column=self.col_dias_inicio + i, value=dia.hc)
            sheet.cell(row=linha, column=self.col_prod, value=semana.produtividade)

            linha_monitoramento = self.get_lin_monitoramento(sheet)
            while sheet.cell(row=linha_monitoramento, column=self.col_nomes).value:
                linha_monitoramento += 1
            sheet.insert_rows(linha_monitoramento)
            sheet.cell(row=linha_monitoramento, column=self.col_nomes, value=projeto.nome)

    def get_planilha_semana_atual(self):
        hoje = date.today()
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        str_data = inicio_semana.strftime('%d.%m')
        return "CATI_Semana_" + str_data

    def get_coluna_diaria(self):
        dia = date.today().weekday()
        if dia <= 5:
            return dia + self.col_nomes + 1
        if dia == 6:
            return dia + self.col_nomes

    def get_projetos_disponiveis(self):
        projetos = []
        sheetname = self.get_planilha_semana_atual()
        sheet = self.workbook[sheetname]
        row = self.lin_nomes_inicio
        while True:
            nome = sheet.cell(row=row, column=self.col_nomes).value
            if not nome:
                break
            hc = sheet.cell(row=row, column=self.col_dia).value
            coleta = sheet.cell(row=row, column=self.col_coleta_total).value
            if coleta is not None and isinstance(hc, (int, float)) and hc > 0:
                projetos.append(nome)
            row += 1
        return list(set(projetos))

    def atualizar_coleta_diaria(self, coletas_por_projeto, hc_por_projeto):
        sheetname = self.get_planilha_semana_atual()
        sheet = self.workbook[sheetname]
        linha = self.get_lin_monitoramento(sheet)
        while True:
            nome = sheet.cell(row=linha, column=self.col_nomes).value
            if not nome:
                break
            if nome in coletas_por_projeto:
                sheet.cell(row=linha, column=self.col_coleta_real, value=coletas_por_projeto[nome])
                sheet.cell(row=linha, column=self.col_hc_real, value=hc_por_projeto[nome])
            linha += 1

    def atualizar_meta_parcial(self):
        dia_para_colunas = {
            0: [6],
            1: [6, 7],
            2: [6, 7, 8],
            3: [6, 7, 8, 9],
            4: [6, 7, 8, 9, 10],
        }
        hoje = date.today().weekday()
        sheet = self.workbook[self.get_planilha_semana_atual()]
        linha = self.lin_nomes_inicio
        while True:
            if not sheet.cell(row=linha, column=self.col_nomes).value:
                break
            if hoje in dia_para_colunas:
                total = sum(
                    sheet.cell(row=linha, column=col).value or 0
                    for col in dia_para_colunas[hoje]
                )
                sheet.cell(row=linha, column=self.col_hc_parcial, value=total)
            linha += 1
        